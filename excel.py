import xlrd
import openpyxl
from openpyxl.utils import get_column_letter
from xlrd import open_workbook
from xlutils.copy import copy
from datetime import datetime, date
from datetime import timedelta


#日期
for days in (0,7):
    # 文件名的时间
    today = ((date.today() - timedelta(days=days)).strftime("%Y%m%d"))
    # 打开execl表
    Mobile01 = xlrd.open_workbook('/Users/wuyuanyuan/Downloads/report/'+ today +'-Mobile-Total-Traffic-Trend.xls')
    Mobile02 = xlrd.open_workbook('/Users/wuyuanyuan/Downloads/report/'+ today + '-Mobile-5catagory-Traffic-Trend.xls')
    Fixed01 = xlrd.open_workbook('/Users/wuyuanyuan/Downloads/report/'+ today + '-Fixed-Total-Traffic-Trend.xls')
    Fixed02 = xlrd.open_workbook('/Users/wuyuanyuan/Downloads/report/'+ today + '-Fixed-5catagory-Traffic-Trend.xls')
    # 定位到第二张表格
    MT = Mobile01.sheet_by_index(1)
    MC = Mobile02.sheet_by_index(1)
    FT = Fixed01.sheet_by_index(1)
    FC = Fixed02.sheet_by_index(1)
    # 打开要写入的excel表
    wb = openpyxl.load_workbook('/Users/wuyuanyuan/Downloads/report/test.xlsx')
    ws1 = wb.get_sheet_by_name('DailyReport')
    if days == 0:       # 如果是当天收到的文件
        column1 = 4         # 应该写在第四列
        ws1.cell(7, 4).value=MT.cell_value(2, 4)     # 写入总流量值
        ws1.cell(12, 4).value=FT.cell_value(2, 4)
        today1 = ((date.today() - timedelta(days=1)).strftime("%Y-%m-%d"))
        ws1.cell(2, 4).value=today1    #写入表头日期
    ## 大类流量
        for row in range(2, 7):
            if MC.cell_value(row, 3) == "Streaming":
                a = MC.cell_value(row, 4)
            elif MC.cell_value(row, 3) == "IM":
                b = MC.cell_value(row, 4)
            elif MC.cell_value(row, 3) == "Game":
                ws1.cell(3, column=column1).value=MC.cell_value(row, 4)
            elif MC.cell_value(row, 3) == "Web_Browsing":
                ws1.cell(4, column=column1).value=MC.cell_value(row, 4)
            elif MC.cell_value(row, 3) == "VoIP":
                ws1.cell(6, column=column1).value=MC.cell_value(row, 4)
        for row in range(2, 7):
            if FC.cell_value(row, 3) == "Streaming":
                c = FC.cell_value(row, 4)
            elif FC.cell_value(row, 3) == "IM":
                d = FC.cell_value(row, 4)
            elif FC.cell_value(row, 3) == "Game":
                ws1.cell(8, column=column1).value=FC.cell_value(row, 4)
            elif FC.cell_value(row, 3) == "Web_Browsing":
                ws1.cell(9, column=column1).value=FC.cell_value(row, 4)
            elif FC.cell_value(row, 3) == "VoIP":
                ws1.cell(11, column=column1).value=FC.cell_value(row, 4)
        ws1.cell(5, column=column1).value=round(float(a)) + round(float(b))
        ws1.cell(10, column=column1).value=round(float(c)) + round(float(d))
    else:
        column1 = 3
        ws1.cell(7, 3).value=MT.cell_value(2, 4)
        ws1.cell(12, 3).value=FT.cell_value(2, 4)
        today7 = ((date.today() - timedelta(days=8)).strftime("%Y-%m-%d"))
        ws1.cell(2, 3).value=today7
    ## 大类流量
        for row in range(2, 7):
            if MC.cell_value(row, 3) == "Streaming":
                a = MC.cell_value(row, 4)
            elif MC.cell_value(row, 3) == "IM":
                b = MC.cell_value(row, 4)
            elif MC.cell_value(row, 3) == "Game":
                ws1.cell(3, column=column1).value=MC.cell_value(row, 4)
            elif MC.cell_value(row, 3) == "Web_Browsing":
                ws1.cell(4, column=column1).value=MC.cell_value(row, 4)
            elif MC.cell_value(row, 3) == "VoIP":
                ws1.cell(6, column=column1).value=MC.cell_value(row, 4)
        for row in range(2, 7):
            if FC.cell_value(row, 3) == "Streaming":
                c = FC.cell_value(row, 4)
            elif FC.cell_value(row, 3) == "IM":
                d = FC.cell_value(row, 4)
            elif FC.cell_value(row, 3) == "Game":
                ws1.cell(8, column=column1).value=FC.cell_value(row, 4)
            elif FC.cell_value(row, 3) == "Web_Browsing":
                ws1.cell(9, column=column1).value=FC.cell_value(row, 4)
            elif FC.cell_value(row, 3) == "VoIP":
                ws1.cell(11, column=column1).value=FC.cell_value(row, 4)
        ws1.cell(5, column=column1).value=round(float(a)) + round(float(b))
        ws1.cell(10, column=column1).value=round(float(c)) + round(float(d))

    wb.save(filename='/Users/wuyuanyuan/Downloads/report/test.xlsx')



#wa = Workbook()
#wr= wa.active
#wr.cell(row=6, column=3).value = MT.cell_value(2, 4)
#wa.save(filename='/Users/wuyuanyuan/Downloads/report/test.xlsx')
